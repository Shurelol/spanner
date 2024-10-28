import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup
import time

# 定义存储结果的列表
schools_info = []
count = 0

for page in range(1, 10):

    # 目标URL
    url = "https://infomap.cdedu.com/Home/Index?all=1&per=1fa7aa19-3fbd-4d9e-86ba-f8f0049943b6" + "&pages=" + str(page)
    # 发起HTTP请求
    response = requests.get(url)
    response.encoding = 'utf-8'  # 设置编码为UTF-8

    # 使用BeautifulSoup解析HTML内容
    soup = BeautifulSoup(response.text, 'html.parser')

    # 查找所有包含学校信息的元素，这里假设每个学校的信息在一个特定的标签中，例如div或li
    schools = soup.find_all('ul', class_='index_ul01')  # 根据实际情况修改选择器

    for school in schools:
        school_li = school.find_all('li')
        for li in school_li:
            school_info = []
            school_info.append(li.find('h1').get_text())
            # 获取其他信息
            for p in li.find_all('p'):
                text = p.get_text().strip()
                if '【学段】' in text:
                    school_info.append(text.replace('【学段】', '').strip())
                elif '【区域】' in text:
                    school_info.append(text.replace('【区域】', '').strip())
                elif '【性质】' in text:
                    school_info.append(text.replace('【性质】', '').strip())
                elif '【电话】' in text:
                    school_info.append(text.replace('【电话】', '').strip())
                elif '【地址】' in text:
                    school_info.append(text.replace('【地址】', '').strip())
                elif '【微信公众号】' in text:
                    school_info.append(text.replace('【微信公众号】', '').strip())
            count += 1
            print(count)
            print(school_info)
            schools_info.append(school_info)

    time.sleep(2)

# 保存
wb = Workbook()
ws = wb.active
ws.title = "高中"

table_titles = ['名称', '学段', '区域', '性质', '电话', '地址', '微信公众号']
title_col = 1
for table_title in table_titles:
    cell = ws.cell(1, title_col)
    cell.value = table_title
    title_col += 1

row = 2
for school_info in schools_info:
    col = 1
    for info in school_info:
        cell = ws.cell(row, col)
        cell.value = info
        col += 1
    row += 1

wb.save(r"E:\\test.xlsx")







